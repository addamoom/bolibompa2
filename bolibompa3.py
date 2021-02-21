import csv
import datetime
import os
from decimal import *
import re
import tkinter as tk
from tkinter import *
from tkinter import filedialog, messagebox, simpledialog, ttk
import openpyxl
from tkinter.ttk import Treeview
from tkinter.ttk import *
import ntpath
from openpyxl import Workbook, styles
from openpyxl.styles import *
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from ttkthemes import ThemedTk

# splashscreenbös
splashscreen = Tk()
splashscreen.overrideredirect(True)
width = splashscreen.winfo_screenwidth()
height = splashscreen.winfo_screenheight()

x_start = (width / 2) - 216
y_start = (height / 2) - 216
splashscreen.geometry('%dx%d+%d+%d' % (432, 432, x_start, y_start))
canvas = Canvas(splashscreen, height=432, width=432, bg="yellow")
canvas.pack()
image = PhotoImage(file="BombermanSigil.gif")

canvas.create_image(216, 216, image=image)

splashscreen.after(2500, splashscreen.destroy)
splashscreen.mainloop()

main_win = ThemedTk(theme="black")
main_win.configure(background='#626262')
main_win.minsize(width=800, height=300)
main_win.title("Bolibompa3")

ttk.Style().configure("black.Treeview", borderwidth=15)

main_win.finale_file = ''
gff_file = ''
create_order = 0

pyrocues = []
dmxques = []
shortcutFile = open('shortcuts.csv', 'r')

plocka_eget = []
plocka_gff = []
errors = []

wb_gff = ''
wb_gff_reset = ''
wb_bulk = openpyxl.load_workbook(filename='Bulklager.xlsx')

ws_gff = ''
ws_gff_reset = ''
ws_bulk = wb_bulk[wb_bulk.sheetnames[0]]

ttk.table = ''
folder_bulk = ''
folder_gff = ''
folder_error = ''

antal_bulk = 0
antal_gff = 0
antal_error = 0

total_bulk = 0
total_gff = 0

enter_factor = ''

analyzed = 0

img_bttn_if = PhotoImage(file="Buttons/button_importera-finale-txt.png")
img_bttn_ig = PhotoImage(file="Buttons/button_importera-gff-prislista.png")
img_bttn_lte = PhotoImage(file="Buttons/button_lagg-till-eltandare.png")
img_bttn_r = PhotoImage(file="Buttons/button_rensa.png")
img_bttn_sfl = PhotoImage(file="Buttons/button_skapa-flammlista.png")
img_bttn_spl = PhotoImage(file="Buttons/button_skapa-plocklista.png")
img_bttn_vp = PhotoImage(file="Buttons/button_visa-pjaser.png")

button_style = Style()
button_style.configure('TButton', bd=0, background='#626262')


def import_finale():
    main_win.finale_file = filedialog.askopenfilename(parent=main_win, initialdir=".", title='Välj Finale-filen')

    main_win.title("Bolibompa3" + " / " + get_file_name(main_win.finale_file))

    global plocka_eget, plocka_gff, errors
    getcontext().prec = 2  # behövs för att tidskonverteringen för flammcuerna ska funka
    dmxques.clear()
    pyrocues.clear()
    plocka_eget = []
    plocka_gff = []
    errors = []

    flag = 0
    with open(main_win.finale_file, newline='', encoding='utf-8') as finalef:
        finale_file = csv.reader(finalef, delimiter='\t')
        for f_row in finale_file:
            if len(f_row) > 2:  # sök inte igenom skabbiga tomma rader
                with open('shortcuts.csv', newline='',
                          encoding='utf-8') as shorts:  # borde testa att flytta ut denna till första open statementet, nu öppnas och stängs filen jätteofta
                    sc = csv.reader(shorts, delimiter=',')
                    for sc_row in sc:
                        if sc_row[0] == f_row[14] and sc_row[1] == f_row[21]:
                            dmxques.append({
                                "tid": f_row[2],
                                "shortcut": sc_row[2],
                                "pos": f_row[14],
                                "effekt": f_row[21]
                            })
                            flag = 1
                            print("Dmxque added")
                            break
            if flag == 0:
                if f_row[21]:
                    #           art.nr              beskrivning       antal       enhetspris summa         kommentar
                    f_cell = f_row[21] + '\t' + f_row[10] + '\t' + '1' + '\t' + '0' + '\t' + '0' + '\t' + ''
                    pyrocues.append(f_cell)  # skulle kunna filtrera bort onödiga saker ur den här arrayen
                    # print("pyrocue added")
                    # print(len(pyrocues))
            flag = 0  # bitches love återställda flaggor
    print('färdig')


def import_gff():
    global gff_file, wb_gff, ws_gff, wb_gff_reset, ws_gff_reset
    gff_file = filedialog.askopenfilename(parent=main_win, initialdir=".", title='Välj GFF-filen')
    wb_gff = openpyxl.load_workbook(gff_file)
    wb_gff_reset = openpyxl.load_workbook(gff_file)
    ws_gff = wb_gff[wb_gff.sheetnames[0]]
    ws_gff_reset = wb_gff_reset[wb_gff_reset.sheetnames[0]]


def write_dmxcues(filepath):
    # prereq: dmxques fylld (finale_import())
    # output: csv-fil med dmxcues formaterade för lightfactory, lägger den i samma dir som finale filen
    fname = os.path.join(filepath, get_file_name(str(main_win.finale_file)))

    with open(re.sub('\.txt$', '', fname) + 'toLightfactory.csv', 'w+', newline='') as csvfile:
        fieldnames = ['namn', 'tid', 'shortcut', '?']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

        for q in dmxques:
            tid = Decimal(q['tid'])
            tidfloor = int(tid)

            frames = round((tid - tidfloor) * 25)
            tidhms = str(datetime.timedelta(seconds=tidfloor))
            tidhms = tidhms + ":" + str(frames)

            writer.writerow({'namn': q['pos'] + " " + q['effekt'], 'tid': tidhms, 'shortcut': q['shortcut'], '?': ''})


def get_file_name(path):
    head, tail = ntpath.split(path)
    return tail


def search_stock(list_of_cues):
    global antal_error
    indices = find_columns(ws_bulk)
    for row_cue in list_of_cues:
        if is_in_bulk(row_cue, indices):
            continue
        elif is_in_gff(row_cue):
            continue
        else:
            row_cue[3] = '0'  # no need to have a price if it's not avaliable for purchase
            row_cue[4] = '0'
            antal_error += 1
            errors.append(row_cue)


# when reading an excel-cell, every value is read as a string, which is why they are parsed

def is_in_bulk(row_cue, _indices):
    global total_bulk, antal_bulk

    for row_bulk in ws_bulk:
        if row_cue[0] == row_bulk[_indices[0]].value and row_bulk[_indices[1]].value > 0:  # check if in stock
            row_bulk[_indices[1]].value = int(row_bulk[_indices[1]].value) - 1  # remove one from stock
            row_cue[3] = float(row_bulk[_indices[2]].value)  # set price
            row_cue[4] = row_cue[3]  # set total price (used later)
            if row_bulk[_indices[3]].value is not None:
                row_cue[5] = row_bulk[_indices[3]].value  # add comment
            plocka_eget.append(row_cue)  # add to plocklista
            total_bulk += row_cue[3]
            antal_bulk += 1
            return True  # go back to searchthingy

    return False


def is_in_gff(row_cue):
    global total_gff, antal_gff
    for row_gff in ws_gff:

        if row_cue[0] == row_gff[3].value:  # if the name matches
            if row_gff[6].value > 0:  # check if in stock
                if row_gff[12].value is None:  # increase order number
                    row_gff[12].value = 0
                elif row_gff[12].value >= row_gff[6].value:        #the current number of ordered pieces has to be lower than the stock
                    return False

                row_cue[3] = float(row_gff[9].value)  # set the price
                row_cue[4] = row_cue[3]  # total price (used later)

                if row_gff[3].value == "900003E":
                    print(row_gff[6].value)
                    print(row_gff[12].value)

                row_gff[12].value += 1

                plocka_gff.append(row_cue)  # add the row to the plocklista
                total_gff += row_cue[3]
                antal_gff += 1
                return True  # go back to searchthingy
            #else:
            #    return False
    return False

#   find_columns loopar igenom den översta raden på bulklagerlistan och letar efter de olika keywordsen.
#   på så vis kan Pjästanten ändra ordning på kolumnerna utan att man måste ändra i Bolibompa
#   om rubrikerna i filen ändras måste de dock ändras här också
def find_columns(_worksheet):
    art_nr = ""
    antal = ""
    pps = ""
    komm = ""
    row_0 = _worksheet[1]
    for i, cell in enumerate(row_0, start=0):
        if row_0[i].value == "Art.nr.":
            art_nr = i
        elif row_0[i].value == "Antal":
            antal = i
        elif row_0[i].value == "Pris/st":
            pps = i
        elif row_0[i].value == "Kommentar":
            komm = i

    return [art_nr, antal, pps, komm]


def sum_list(lista):
    lista.sort()
    summed_list = []
    index = 0
    summed_list.append(lista[0])  # add the first element of the old to the new
    lista.pop(0)  # and remove it from the old
    for row in lista:  # loop through the old list
        row[2] = int(row[2])
        if row[0] == summed_list[index][0]:
            summed_list[index][2] = int(summed_list[index][2]) + 1  # increment antal
            summed_list[index][4] = float(summed_list[index][4]) + float(summed_list[index][3])  # öka summa
        else:
            summed_list[index][4] = round(float(summed_list[index][4]), 2)  # runda av summan
            summed_list.append(row)  # lägg till nya raden
            index += 1

    return summed_list


def kbk_pyro():
    if analyzed:
        varning = messagebox.askyesno("Varning", "Du kommer att skriva över filer nu, vill du fortsätta?")
        if varning:
            location = filedialog.askdirectory()
            print_list(location)
            path = os.path.join(location, 'GFF_Order.xlsx')
            currentrow = 1
            for eachRow in ws_gff_reset.iter_rows():
                ws_gff.cell(row=currentrow, column=7).value = ws_gff_reset.cell(row=currentrow, column=7).value
                currentrow += 1

            wb_gff.save(path)
            wb_bulk.save('Bulklager.xlsx')

            messagebox.showinfo("Färdigt!", "Nu finns det listor. Coolt va?")
    else:
        messagebox.showinfo("Fel!", 'Klicka på "Visa Lista" först  , din smurf!')


def kbk_flames():
    if dmxques:
        path = filedialog.askdirectory()
        write_dmxcues(path)
        messagebox.showinfo("Spännande!", "Flammfilen finns nu ")


def scan_list():
    global table, analyzed, ign_1m, ign_5m, ign_old, plocka_eget, plocka_gff, errors
    if main_win.finale_file and gff_file:
        print('Båda filerna finns')
        # search_assortment()

        pyrocues.pop(0)
        pcs = csv.reader(pyrocues, delimiter='\t')
        search_stock(pcs)

        if plocka_eget:
            plocka_eget = sum_list(plocka_eget)
            display_lists(folder_bulk, plocka_eget)
            table.item('folder_bulk', values=['', antal_bulk, '', round(total_bulk, 2)])
        if plocka_gff:
            plocka_gff = sum_list(plocka_gff)
            display_lists(folder_gff, plocka_gff)
            table.item('folder_gff', values=['', antal_gff, '', round(total_gff, 2)])
        if errors:
            errors = sum_list(errors)
            display_lists(folder_error, errors)
            table.item('folder_error', values=['', antal_error, '', ''])
        analyzed = 1
    else:
        messagebox.showinfo("Varning!", "Du måste välja filer först")


def display_lists(folder, list):
    for row in list:
        print(row)
        table.insert(folder, "end", text=row[0], values=[row[1], row[2], row[3], row[4], row[5]])


def print_list(location):
    wb1 = Workbook()
    path = os.path.join(location, 'plocklista.xlsx')

    ws1 = wb1.active
    ws1.title = 'Pjäser'
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 60
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 10
    ws1.column_dimensions['E'].width = 10
    ws1.column_dimensions['F'].width = 30

    ws1.append(['Art.nr', 'Beskrivning', 'Antal', 'Pris', 'Summa', 'Kommentar'])

    ws1.append([''])
    ws1.append(['Från Bulk'])
    if plocka_eget:
        for row in plocka_eget:
            ws1.append(row)

    ws1.append([''])
    ws1.append(['Från GFF'])
    if plocka_gff:
        for row in plocka_gff:
            ws1.append(row)

    for row in ws1:
        if row[0].value == 'Från Bulk' or row[0].value == 'Från GFF':
            for i in range(0, 5):
                row[i].font = Font(name='Calibri', bold=True)
                row[i].fill = PatternFill(fill_type=SOLID, fgColor='00F6A54A')
        else:
            row[2].fill = PatternFill(fill_type=SOLID, fgColor='00cfd8dc')

    for cell in ws1["1:1"]:
        cell.font = Font(name='Calibri', bold=True)
        cell.fill = PatternFill(fill_type=SOLID, fgColor='00F57D00')

    wb1.save(filename=path)


def add_ign():
    global total_bulk, ws_bulk, antal_bulk
    ign_1m = simpledialog.askinteger(title="", prompt="Hur många eltändare 1m (Svart)?", initialvalue=0)
    ign_5m = simpledialog.askinteger(title="", prompt="Hur många eltändare 5m  (Orange)?", initialvalue=0)

    i1m_price = 0
    i5m_price = 0

    for row in ws_bulk:
        if row[0].value == 'P-IGN-1M':
            row[3].value -= ign_1m
            i1m_price = float(row[6].value)
        elif row[0].value == 'P-IGN-5M':
            row[3].value -= ign_5m
            i5m_price = float(row[6].value)

    igniters = []

    # Varning för fulkod. Känsliga programmerare bör blunda
    if ign_1m > 0:
        antal_bulk += ign_1m
        igniters.append(
            ['P-IGN-1M'] + ['Eltändare 1m (svart)'] + [ign_1m] + [i1m_price] + [round(i1m_price * ign_1m, 2)] + [''])

    if ign_5m > 0:
        antal_bulk += ign_5m
        igniters.append(
            ['P-IGN-5M'] + ['Eltändare 5m (Orange)'] + [ign_5m] + [i5m_price] + [round(i5m_price * ign_5m, 2)] + [''])

    if igniters:
        total_bulk += ((i1m_price * ign_1m) + (i5m_price * ign_5m))
        total_bulk = round(total_bulk, 2)
        # search_stock(igniters)
        display_lists(folder_bulk, igniters)
        plocka_eget.extend(igniters)
        table.item('folder_bulk', values=['', antal_bulk, '', total_bulk, ''])


def re_init():
    global table, folder_bulk, folder_gff, folder_error, total_bulk, total_gff, pyrocues, dmxques, shortcutFile
    global plocka_eget, plocka_gff, errors, wb_gff, ws_gff, ws_bulk, analyzed, wb_bulk, ign_1m, ign_5m
    global antal_bulk, antal_error, antal_gff

    main_win.title("Bolibompa3")

    main_win.finale_file = ''
    main_win.gff_file = ''

    total_gff = 0
    total_bulk = 0

    antal_bulk = 0
    antal_error = 0
    antal_gff = 0

    table.delete(*table.get_children())
    folder_bulk = table.insert("", 1, 'folder_bulk', text="Bulklager", values=['', '', '', total_bulk, ''],
                               tags='folder')
    folder_gff = table.insert("", 2, 'folder_gff', text="GFF", values=['', '', '', total_gff, ''], tags='folder')
    folder_error = table.insert("", 3, 'folder_error', text="Error", tags='folder')

    table.tag_configure('folder', font='bold')


    pyrocues = []
    plocka_eget = []
    plocka_gff = []
    errors = []

    display_lists(folder_bulk, plocka_eget)
    display_lists(folder_gff, plocka_gff)
    display_lists(folder_error, errors)

    shortcutFile = open('shortcuts.csv', 'r')

    #wb_gff.close()
    wb_gff = ''
    wb_bulk = openpyxl.load_workbook(filename='Bulklager.xlsx')

    ws_gff = ''
    ws_bulk = wb_bulk[wb_bulk.sheetnames[0]]

    analyzed = 0

    ign_1m = 0
    ign_5m = 0

    display_lists(folder_bulk, plocka_eget)
    display_lists(folder_gff, plocka_gff)
    display_lists(folder_error, errors)


    display_lists()

    messagebox.showinfo("Klar!", "Din session är nu rensad. Var god välj nya filer")


def init(main_win):
    global table, folder_bulk, folder_gff, folder_error, enter_factor
    global img_bttn_if, img_bttn_ig, img_bttn_lte, img_bttn_r, img_bttn_sfl, img_bttn_spl, img_bttn_vp
    info_frame = tk.Frame(main_win, bd=0)
    info_frame.grid(column=0, row=0, sticky="ns")
    info_frame.pack(fill='both', expand=TRUE)

    #  info_scroll = Scrollbar(info_frame)
    #  info_scroll.pack(side=RIGHT)
    table_style = ttk.Style()
    table_style.layout('Pyrot.Treeview', [('Pyrot.Treeview.treearea', {'sticky': 'nswe'})])
    table = ttk.Treeview(info_frame, style='Pyrot.Treeview')  # gör denna global

    table["columns"] = ("one", "two", "three", "four", "five", "six")
    table.column("#0", width=150, minwidth=150, stretch=NO)
    table.column("#1", width=300, minwidth=200)
    table.column("#2", width=50, minwidth=50, stretch=NO)
    table.column("#3", width=50, minwidth=50, stretch=NO)
    table.column("#4", width=75, minwidth=75, stretch=NO)
    table.column("#5", width=500, minwidth=200)

    table.heading("#0", text="Art. NR", anchor=W)
    table.heading("#1", text="Beskrivning", anchor=W)
    table.heading("#2", text="Antal", anchor=W)
    table.heading("#3", text="Pris", anchor=W)
    table.heading("#4", text="Summa", anchor=W)
    table.heading("#5", text="Kommentar", anchor=W)
    # Level 1
    folder_bulk = table.insert("", 1, 'folder_bulk', text="Bulklager", values=['', '', '', total_bulk, ''],
                               tags='folder')
    folder_gff = table.insert("", 2, 'folder_gff', text="GFF", values=['', '', '', total_gff, ''], tags='folder')
    folder_error = table.insert("", 3, 'folder_error', text="Error", tags='folder')

    table.tag_configure('folder', font='bold')

    table.pack(side=TOP, fill='y', expand=True)  # , fill=X)

    button_frame = tk.Frame(main_win, bg='#626262')

    bttn_add_ign = tk.Button(button_frame, image=img_bttn_lte, command=add_ign, bd=0, bg='#626262',
                             activebackground='#626262', highlightthickness=0)
    bttn_add_ign.pack(side=TOP)

    bttn_import_finale = tk.Button(button_frame, image=img_bttn_if, command=import_finale, bd=0, bg='#626262',
                                   activebackground='#626262', highlightthickness=0)
    bttn_import_finale.pack(side=LEFT)

    bttn_import_gff = tk.Button(button_frame, image=img_bttn_ig, command=import_gff, bd=0, bg='#626262',
                                activebackground='#626262', highlightthickness=0)
    bttn_import_gff.pack(side=LEFT)

    bttn_search_list = tk.Button(button_frame, image=img_bttn_vp, command=scan_list, bd=0, bg='#626262',
                                 activebackground='#626262', highlightthickness=0)
    bttn_search_list.pack(side=LEFT)

    bttn_transact = tk.Button(button_frame, image=img_bttn_spl, command=kbk_pyro, bd=0, bg='#626262',
                              activebackground='#626262', highlightthickness=0)
    bttn_transact.pack(side=RIGHT)

    bttn_flames = tk.Button(button_frame, image=img_bttn_sfl, command=kbk_flames, bd=0, bg='#626262',
                            activebackground='#626262', highlightthickness=0)
    bttn_flames.pack(side=RIGHT)

    bttn_clear = tk.Button(button_frame, image=img_bttn_r, command=re_init, bd=0, bg='#626262',
                           activebackground='#626262', highlightthickness=0)
    bttn_clear.pack(side=RIGHT)

    info_frame.pack(side=TOP, fill=Y)
    button_frame.pack(side=BOTTOM)


init(main_win)

main_win.mainloop()
